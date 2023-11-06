import sqlite3


class DBModule:
    connection = sqlite3.connect("users.db")
    cursor = connection.cursor()

    @classmethod
    def create_table(cls):
        cls.cursor.execute('''CREATE TABLE IF NOT EXISTS cars (
            id INTEGER PRIMARY KEY,
            make TEXT,
            year INTEGER,
            color TEXT,
            country TEXT
        )''')

        cls.connection.commit()

    @classmethod
    def insert_car(cls, make, year, color, country):
        cls.cursor.execute("INSERT INTO cars (make, year, color, country) VALUES (?, ?, ?, ?)",
                           (make, year, color, country))
        cls.connection.commit()

    @classmethod
    def get_cars_by_color(cls, color):
        cls.cursor.execute("SELECT make, year, country FROM cars WHERE color = ?", (color,))
        return cls.cursor.fetchall()

    @classmethod
    def export_to_excel(cls, filename):
        import openpyxl

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Cars"

        cls.cursor.execute("SELECT make, year, color, country FROM cars")
        cars = cls.cursor.fetchall()

        for row, car in enumerate(cars, start=2):
            worksheet.cell(row=row, column=1, value=car[0])
            worksheet.cell(row=row, column=2, value=car[1])
            worksheet.cell(row=row, column=3, value=car[2])
            worksheet.cell(row=row, column=4, value=car[3])

        workbook.save(filename)

DBModule.create_table()
def db_data_check():
    conn = sqlite3.connect('users.db')
    cur = conn.cursor()
    cur.execute('SELECT * FROM cars')
    data_all = cur.fetchall()
    for data_one in data_all:
        print(data_one)
    conn.close()


db_data_check()
